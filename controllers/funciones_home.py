from werkzeug.utils import secure_filename
import uuid
from conexion.conexionBD import *  # Conexión a BD
import datetime
import re
import os
from os import remove  # Modulo  para remover archivo
from os import path  # Modulo para obtener la ruta o directorio
import openpyxl  # Para generar el excel
from flask import send_file
import requests
import base64

def procesar_form_empleado(dataForm, foto_perfil):
    # Formateando Salario
    salario_sin_puntos = re.sub('[^0-9]+', '', dataForm['salario_empleado'])
    # convertir salario a INT
    salario_entero = int(salario_sin_puntos)

    result_foto_perfil = procesar_imagen_perfil(foto_perfil)
    try:
        with connectionBD_railway() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:

                sql = "INSERT INTO tbl_empleados (nombre_empleado, apellido_empleado, sexo_empleado, telefono_empleado, email_empleado, profesion_empleado, foto_empleado, salario_empleado) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)"

                # Creando una tupla con los valores del INSERT
                valores = (dataForm['nombre_empleado'], dataForm['apellido_empleado'], dataForm['sexo_empleado'],
                            dataForm['telefono_empleado'], dataForm['email_empleado'], dataForm['profesion_empleado'], result_foto_perfil, salario_entero)
                cursor.execute(sql, valores)

                conexion_MySQLdb.commit()
                resultado_insert = cursor.rowcount
                return resultado_insert

    except Exception as e:
        return f'Se produjo un error en procesar_form_empleado: {str(e)}'


def procesar_imagen_perfil(foto):
    try:
        # Nombre original del archivo
        filename = secure_filename(foto.filename)
        extension = os.path.splitext(filename)[1]

        # Creando un string de 50 caracteres
        nuevoNameFile = (uuid.uuid4().hex + uuid.uuid4().hex)[:100]
        nombreFile = nuevoNameFile + extension

        # Construir la ruta completa de subida del archivo
        basepath = os.path.abspath(os.path.dirname(__file__))
        upload_dir = os.path.join(basepath, f'../static/fotos_empleados/')

        # Validar si existe la ruta y crearla si no existe
        if not os.path.exists(upload_dir):
            os.makedirs(upload_dir)
            # Dando permiso a la carpeta
            os.chmod(upload_dir, 0o755)

        # Construir la ruta completa de subida del archivo
        upload_path = os.path.join(upload_dir, nombreFile)
        foto.save(upload_path)

        return nombreFile

    except Exception as e:
        print("Error al procesar archivo:", e)
        return []


# Lista de Empleados
def sql_lista_empleadosBD():
    try:
        with connectionBD_railway() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = (f"""
                    SELECT 
                        e.id_empleado,
                        e.nombre_empleado, 
                        e.apellido_empleado,
                        e.salario_empleado,
                        e.foto_empleado,
                        CASE
                            WHEN e.sexo_empleado = 1 THEN 'Masculino'
                            ELSE 'Femenino'
                        END AS sexo_empleado
                    FROM tbl_empleados AS e
                    ORDER BY e.id_empleado DESC
                    """)
                cursor.execute(querySQL,)
                empleadosBD = cursor.fetchall()
        return empleadosBD
    except Exception as e:
        print(
            f"Errro en la función sql_lista_empleadosBD: {e}")
        return None

# Lista de Empleados
def sql_lista_inventariobodegaBD():
    try:
        with connectionBD_inv() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = (f"""
                    SELECT  e.creation_date, e.Bodega, e.Material, e.Subproducto, e.CantidadDisponible, e.Ubicacion 
                    FROM inventario_bodega e
                    order by e.CantidadDisponible desc limit 10;
                    """)
                cursor.execute(querySQL,)
                inventarioBD = cursor.fetchall()
        return inventarioBD
    except Exception as e:
        print(
            f"Error en la función sql_lista_inventariobodegaBD: {e}")
        return []

def sql_lista_inventariobodegaBD_oms():
    try:
        with connectionBD_inv() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = (f"""
                    SELECT  e.creation_date, e.Sku, e.NombreSku, e.Cantidad
                    FROM inventario_OMS e
                    where e.NombreSku NOT LIKE '%PRESTAMO%' and e.NombreSku NOT LIKE '%REMOTO%' and e.NombreSku NOT LIKE '%SIM%'
                    order by e.Cantidad desc limit 10;
                    """)
                cursor.execute(querySQL,)
                inventarioBD_oms = cursor.fetchall()
        return inventarioBD_oms
    except Exception as e:
        print(
            f"Error en la función sql_lista_inventariobodegaBD_oms: {e}")
        return []

def sql_lista_tokenx():
    try:
        with connectionBD_railway() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = (f"""
                    SELECT * FROM token_poliedro ORDER BY `fecha_registro` desc LIMIT 10;
                    """)
                cursor.execute(querySQL,)
                token = cursor.fetchall()
        return token
    except Exception as e:
        print(
            f"Error en la función sql_lista_token: {e}")
        return None
def sql_lista_token(tipo_token, campaing):
    try:
        with connectionBD_railway() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                querySQL = ("""
                        SELECT t.token,t.fecha_registro  
                        FROM token_poliedro as t
                        WHERE t.tipo_token LIKE %s AND t.campaing LIKE %s
                        ORDER BY t.fecha_registro desc LIMIT 5
                    """)
                search_bodega_pattern = f"%{tipo_token}%"  # Para tipo token
                search_producto_pattern = f"%{campaing}%"  # Para campaña
                mycursor.execute(querySQL, (search_bodega_pattern, search_producto_pattern))
                resultadobusquedainv_bodega_pro = mycursor.fetchall()
                return resultadobusquedainv_bodega_pro

    except Exception as e:
        print(f"Error en la función sql_lista_token: {e}")
        return None

# Detalles del Empleado
def sql_detalles_empleadosBD(idEmpleado):
    try:
        with connectionBD_railway() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = ("""
                    SELECT 
                        e.id_empleado,
                        e.nombre_empleado, 
                        e.apellido_empleado,
                        e.salario_empleado,
                        CASE
                            WHEN e.sexo_empleado = 1 THEN 'Masculino'
                            ELSE 'Femenino'
                        END AS sexo_empleado,
                        e.telefono_empleado, 
                        e.email_empleado,
                        e.profesion_empleado,
                        e.foto_empleado,
                        DATE_FORMAT(e.fecha_registro, '%Y-%m-%d %h:%i %p') AS fecha_registro
                    FROM tbl_empleados AS e
                    WHERE id_empleado =%s
                    ORDER BY e.id_empleado DESC
                    """)
                cursor.execute(querySQL, (idEmpleado,))
                empleadosBD = cursor.fetchone()
        return empleadosBD
    except Exception as e:
        print(
            f"Errro en la función sql_detalles_empleadosBD: {e}")
        return None


# Funcion Empleados Informe (Reporte)
def empleadosReporte():
    try:
        with connectionBD_railway() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = ("""
                    SELECT 
                        e.id_empleado,
                        e.nombre_empleado, 
                        e.apellido_empleado,
                        e.salario_empleado,
                        e.email_empleado,
                        e.telefono_empleado,
                        e.profesion_empleado,
                        DATE_FORMAT(e.fecha_registro, '%d de %b %Y %h:%i %p') AS fecha_registro,
                        CASE
                            WHEN e.sexo_empleado = 1 THEN 'Masculino'
                            ELSE 'Femenino'
                        END AS sexo_empleado
                    FROM tbl_empleados AS e
                    ORDER BY e.id_empleado DESC
                    """)
                cursor.execute(querySQL,)
                empleadosBD = cursor.fetchall()
        return empleadosBD
    except Exception as e:
        print(
            f"Errro en la función empleadosReporte: {e}")
        return None

def whatReporte():
    try:
        with connectionBD_railway() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = ("""
                    SELECT 
                        e.id,
                        e.first_name, 
                        e.phone,
                        DATE_FORMAT(STR_TO_DATE(e.created_at, '%Y-%m-%dT%H:%i:%s.%fZ'), '%Y-%m-%d %H:%i:%s') AS fecha,
                        e.nameLabels
                    FROM Historial_Whatsapp AS e
                    ORDER BY DATE(e.created_at) DESC
                    """)
                cursor.execute(querySQL,)
                empleadosBD = cursor.fetchall()
        return empleadosBD
    except Exception as e:
        print(
            f"Errro en la función empleadosReporte: {e}")
        return None

def generarReporteExcel():
    dataEmpleados = empleadosReporte()
    wb = openpyxl.Workbook()
    hoja = wb.active

    # Agregar la fila de encabezado con los títulos
    cabeceraExcel = ("Nombre", "Apellido", "Sexo",
                     "Telefono", "Email", "Profesión", "Salario", "Fecha de Ingreso")

    hoja.append(cabeceraExcel)

    # Formato para números en moneda colombiana y sin decimales
    formato_moneda_colombiana = '#,##0'

    # Agregar los registros a la hoja
    for registro in dataEmpleados:
        nombre_empleado = registro['nombre_empleado']
        apellido_empleado = registro['apellido_empleado']
        sexo_empleado = registro['sexo_empleado']
        telefono_empleado = registro['telefono_empleado']
        email_empleado = registro['email_empleado']
        profesion_empleado = registro['profesion_empleado']
        salario_empleado = registro['salario_empleado']
        fecha_registro = registro['fecha_registro']

        # Agregar los valores a la hoja
        hoja.append((nombre_empleado, apellido_empleado, sexo_empleado, telefono_empleado, email_empleado, profesion_empleado,
                     salario_empleado, fecha_registro))

        # Itera a través de las filas y aplica el formato a la columna G
        for fila_num in range(2, hoja.max_row + 1):
            columna = 7  # Columna G
            celda = hoja.cell(row=fila_num, column=columna)
            celda.number_format = formato_moneda_colombiana

    fecha_actual = datetime.datetime.now()
    archivoExcel = f"Reporte_empleados_{fecha_actual.strftime('%Y_%m_%d')}.xlsx"
    carpeta_descarga = "../static/downloads-excel"
    ruta_descarga = os.path.join(os.path.dirname(
        os.path.abspath(__file__)), carpeta_descarga)

    if not os.path.exists(ruta_descarga):
        os.makedirs(ruta_descarga)
        # Dando permisos a la carpeta
        os.chmod(ruta_descarga, 0o755)

    ruta_archivo = os.path.join(ruta_descarga, archivoExcel)
    wb.save(ruta_archivo)

    # Enviar el archivo como respuesta HTTP
    return send_file(ruta_archivo, as_attachment=True)


def generarReporteExcelwhat():
    dataEmpleados = whatReporte()
    wb = openpyxl.Workbook()
    hoja = wb.active

    # Agregar la fila de encabezado con los títulos
    cabeceraExcel = ("Id", "Nombre Cliente", "Telefono","Fecha de Ingreso","Etiquetas")

    hoja.append(cabeceraExcel)

    # Agregar los registros a la hoja
    for registro in dataEmpleados:
        Id = registro['Id']
        Nombre_Cliente = registro['Nombre Cliente']
        Telefono = registro['Telefono']
        Fecha_de_Ingreso = registro['Fecha de Ingreso']
        Etiquetas = registro['Etiquetas']

        # Agregar los valores a la hoja
        hoja.append((Id, Nombre_Cliente, Telefono, Fecha_de_Ingreso, Etiquetas))

    fecha_actual = datetime.datetime.now()
    archivoExcel = f"Reporte_whatsapp_{fecha_actual.strftime('%Y_%m_%d')}.xlsx"
    carpeta_descarga = "../static/downloads-excel"
    ruta_descarga = os.path.join(os.path.dirname(
        os.path.abspath(__file__)), carpeta_descarga)

    if not os.path.exists(ruta_descarga):
        os.makedirs(ruta_descarga)
        # Dando permisos a la carpeta
        os.chmod(ruta_descarga, 0o755)

    ruta_archivo = os.path.join(ruta_descarga, archivoExcel)
    wb.save(ruta_archivo)

    # Enviar el archivo como respuesta HTTP
    return send_file(ruta_archivo, as_attachment=True)

def buscarEmpleadoBD(search):
    try:
        with connectionBD_railway() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                querySQL = ("""
                        SELECT 
                            e.id_empleado,
                            e.nombre_empleado, 
                            e.apellido_empleado,
                            e.salario_empleado,
                            CASE
                                WHEN e.sexo_empleado = 1 THEN 'Masculino'
                                ELSE 'Femenino'
                            END AS sexo_empleado
                        FROM tbl_empleados AS e
                        WHERE e.nombre_empleado LIKE %s 
                        ORDER BY e.id_empleado DESC
                    """)
                search_pattern = f"%{search}%"  # Agregar "%" alrededor del término de búsqueda
                mycursor.execute(querySQL, (search_pattern,))
                resultado_busqueda = mycursor.fetchall()
                return resultado_busqueda

    except Exception as e:
        print(f"Ocurrió un error en def buscarEmpleadoBD: {e}")
        return []


def buscarInventarioBD(search):
    try:
        with connectionBD_inv() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                querySQL = ("""
                        SELECT  e.creation_date, e.Bodega, e.Material, e.Subproducto, e.CantidadDisponible, e.Ubicacion 
                        FROM inventario_bodega AS e
                        WHERE e.Subproducto LIKE %s 
                        ORDER BY e.Subproducto asc,e.CantidadDisponible asc LIMIT 20
                    """)
                search_pattern = f"%{search}%"  # Agregar "%" alrededor del término de búsqueda
                mycursor.execute(querySQL, (search_pattern,))
                resultadobusquedainv = mycursor.fetchall()
                return resultadobusquedainv

    except Exception as e:
        print(f"Ocurrió un error en def buscarInventarioBD: {e}")
        return []
    
def buscarInventarioBD_bodega(search):
    try:
        with connectionBD_inv() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                querySQL = ("""
                        SELECT  e.creation_date, e.Bodega, e.Material, e.Subproducto, e.CantidadDisponible, e.Ubicacion 
                        FROM inventario_bodega AS e
                        WHERE e.Bodega LIKE %s 
                        ORDER BY e.Subproducto asc,e.CantidadDisponible asc LIMIT 20
                    """)
                search_pattern = f"%{search}%"  # Agregar "%" alrededor del término de búsqueda
                mycursor.execute(querySQL, (search_pattern,))
                resultadobusquedainv_bodega = mycursor.fetchall()
                return resultadobusquedainv_bodega

    except Exception as e:
        print(f"Ocurrió un error en def buscarInventarioBD: {e}")
        return []
    
def buscarInventarioBD_bodega_pro(search_bodega, search_producto):
    try:
        with connectionBD_inv() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                querySQL = ("""
                        SELECT  e.creation_date, e.Bodega, e.Material, e.Subproducto, e.CantidadDisponible, e.Ubicacion 
                        FROM inventario_bodega AS e
                        WHERE e.Bodega LIKE %s AND e.Subproducto LIKE %s
                        ORDER BY e.Subproducto asc,e.CantidadDisponible asc LIMIT 20
                    """)
                search_bodega_pattern = f"%{search_bodega}%"  # Para Bodega
                search_producto_pattern = f"%{search_producto}%"  # Para Producto
                mycursor.execute(querySQL, (search_bodega_pattern, search_producto_pattern))
                resultadobusquedainv_bodega_pro = mycursor.fetchall()
                return resultadobusquedainv_bodega_pro

    except Exception as e:
        print(f"Ocurrió un error en def buscarInventarioBD_bodega_pro: {e}")
        return []

def buscarInventarioBD_bodega_oms(search_producto):
    try:
        with connectionBD_inv() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                querySQL = ("""
                        SELECT  e.creation_date, e.Sku, e.NombreSku, e.Cantidad
                        FROM inventario_OMS e
                        WHERE e.NombreSku LIKE %s
                        order by e.NombreSku asc,e.Cantidad asc limit 20
                    """)
                
                search_producto_pattern = f"%{search_producto}%"
                mycursor.execute(querySQL, (search_producto_pattern,))
                resultadobusquedainv_oms = mycursor.fetchall()
                return resultadobusquedainv_oms

    except Exception as e:
        print(f"Ocurrió un error en def buscarInventarioBD_bodega_oms: {e}")
        return []

def buscarEmpleadoUnico(id):
    try:
        with connectionBD_railway() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                querySQL = ("""
                        SELECT 
                            e.id_empleado,
                            e.nombre_empleado, 
                            e.apellido_empleado,
                            e.sexo_empleado,
                            e.telefono_empleado,
                            e.email_empleado,
                            e.profesion_empleado,
                            e.salario_empleado,
                            e.foto_empleado
                        FROM tbl_empleados AS e
                        WHERE e.id_empleado =%s LIMIT 1
                    """)
                mycursor.execute(querySQL, (id,))
                empleado = mycursor.fetchone()
                return empleado

    except Exception as e:
        print(f"Ocurrió un error en def buscarEmpleadoUnico: {e}")
        return []


def procesar_actualizacion_form(data):
    try:
        with connectionBD_railway() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                nombre_empleado = data.form['nombre_empleado']
                apellido_empleado = data.form['apellido_empleado']
                sexo_empleado = data.form['sexo_empleado']
                telefono_empleado = data.form['telefono_empleado']
                email_empleado = data.form['email_empleado']
                profesion_empleado = data.form['profesion_empleado']

                salario_sin_puntos = re.sub(
                    '[^0-9]+', '', data.form['salario_empleado'])
                salario_empleado = int(salario_sin_puntos)
                id_empleado = data.form['id_empleado']

                if data.files['foto_empleado']:
                    file = data.files['foto_empleado']
                    fotoForm = procesar_imagen_perfil(file)

                    querySQL = """
                        UPDATE tbl_empleados
                        SET 
                            nombre_empleado = %s,
                            apellido_empleado = %s,
                            sexo_empleado = %s,
                            telefono_empleado = %s,
                            email_empleado = %s,
                            profesion_empleado = %s,
                            salario_empleado = %s,
                            foto_empleado = %s
                        WHERE id_empleado = %s
                    """
                    values = (nombre_empleado, apellido_empleado, sexo_empleado,
                              telefono_empleado, email_empleado, profesion_empleado,
                              salario_empleado, fotoForm, id_empleado)
                else:
                    querySQL = """
                        UPDATE tbl_empleados
                        SET 
                            nombre_empleado = %s,
                            apellido_empleado = %s,
                            sexo_empleado = %s,
                            telefono_empleado = %s,
                            email_empleado = %s,
                            profesion_empleado = %s,
                            salario_empleado = %s
                        WHERE id_empleado = %s
                    """
                    values = (nombre_empleado, apellido_empleado, sexo_empleado,
                              telefono_empleado, email_empleado, profesion_empleado,
                              salario_empleado, id_empleado)

                cursor.execute(querySQL, values)
                conexion_MySQLdb.commit()

        return cursor.rowcount or []
    except Exception as e:
        print(f"Ocurrió un error en procesar_actualizacion_form: {e}")
        return None


# Lista de Usuarios creados
def lista_usuariosBD():
    try:
        with connectionBD_railway() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "SELECT id, name_surname, email_user, rol, token, campaing, created_user FROM users"
                cursor.execute(querySQL,)
                usuariosBD = cursor.fetchall()
        return usuariosBD
    except Exception as e:
        print(f"Error en lista_usuariosBD : {e}")
        return []

# Lista de equipos creados
def lista_equiposBD():
    try:
        with connectionBD_railway() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "SELECT id,tipo_equipo,marca,producto,gamma FROM inventario_digital"
                cursor.execute(querySQL,)
                equiposBD = cursor.fetchall()
        return equiposBD
    except Exception as e:
        print(f"Error en lista_usuariosBD : {e}")
        return []
    
# buscar equipos creados
def buscador_lista_equiposBD(search_producto):
    try:
        with connectionBD_inv() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                querySQL = ("""
                        SELECT tipo_equipo,marca,producto,gamma FROM inventario_digital
                        WHERE marca LIKE %s or producto LIKE %s
                        limit 20
                    """)
                
                search_producto_pattern = f"%{search_producto}%"
                mycursor.execute(querySQL, (search_producto_pattern,))
                resultadobusquedainv_oms = mycursor.fetchall()
                return resultadobusquedainv_oms

    except Exception as e:
        print(f"Ocurrió un error en def buscador_lista_equiposBD: {e}")
        return []

# Eliminar Empleado
def eliminarEmpleado(id_empleado, foto_empleado):
    try:
        with connectionBD_railway() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "DELETE FROM tbl_empleados WHERE id_empleado=%s"
                cursor.execute(querySQL, (id_empleado,))
                conexion_MySQLdb.commit()
                resultado_eliminar = cursor.rowcount

                if resultado_eliminar:
                    # Eliminadon foto_empleado desde el directorio
                    basepath = path.dirname(__file__)
                    url_File = path.join(
                        basepath, '../static/fotos_empleados', foto_empleado)

                    if path.exists(url_File):
                        remove(url_File)  # Borrar foto desde la carpeta

        return resultado_eliminar
    except Exception as e:
        print(f"Error en eliminarEmpleado : {e}")
        return []


# Eliminar usuario
def eliminarUsuario(id):
    try:
        with connectionBD_railway() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "DELETE FROM users WHERE id=%s"
                cursor.execute(querySQL, (id,))
                conexion_MySQLdb.commit()
                resultado_eliminar = cursor.rowcount

        return resultado_eliminar
    except Exception as e:
        print(f"Error en eliminarUsuario : {e}")
        return []
    
# Eliminar equipo digital
def eliminarequipo(id):
    try:
        with connectionBD_railway() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "DELETE FROM inventario_digital WHERE id=%s"
                cursor.execute(querySQL, (id,))
                conexion_MySQLdb.commit()
                resultado_eliminar = cursor.rowcount

        return resultado_eliminar
    except Exception as e:
        print(f"Error en eliminarUsuario : {e}")
        return []

def guardar_url(id, url_larga, url_corta):
    try:
        with connectionBD_railway() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "UPDATE inventario_digital SET url_larga='%s',url_corta='%s' WHERE id=%s;"
                cursor.execute(querySQL, (url_larga,url_corta,id,))
                conexion_MySQLdb.commit()
                resultado_agregar_url_corta = cursor.rowcount

        return resultado_agregar_url_corta
    except Exception as e:
        print(f"Error modificando url´s cortas : {e}")
        return []

def enviar_sms(numero, mensaje):
    
    usuario = "api.pdzlr"
    password = ".NQlu2t_GISxYWB@r3w1N2jsha3V.P"
    credenciales = f"{usuario}:{password}"
    token_base64 = base64.b64encode(credenciales.encode()).decode()

    #print(f"Authorization: Basic {token_base64}")
    
    url = "https://api-sms.masivapp.com/send-message"
    
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Basic {token_base64}"
    }
    
    data = {
        "to": f'57{numero}',
        "text": mensaje,
        "customData": "Envio inventario digital"
    }
    print(data)
    try:
        response = requests.post(url, headers=headers, json=data)
        print(f"📥 Respuesta de MasivApp: {response.status_code} - {response.text}")
        return response.json()
    except Exception as e:
        print(f"❌ Error al enviar SMS: {str(e)}")
        return {"error": "Error interno al enviar el SMS"}
